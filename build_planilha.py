from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Gerenciador de CRM"

FONT_NAME = "Arial"
header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="1E3A8A")
section_font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
section_fill = PatternFill("solid", fgColor="2563EB")
sub_font = Font(name=FONT_NAME, bold=True, color="0F1B2D", size=10)
sub_fill = PatternFill("solid", fgColor="DBEAFE")
body_font = Font(name=FONT_NAME, size=10)
note_fill = PatternFill("solid", fgColor="FEF3C7")
thin = Side(border_style="thin", color="D1D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

HEADERS = ["Métrica", "Dado ou Função", "Se é função, qual?", "Tem derivação?", "Local (Fonte de Dados)", "Output / Objetivo"]

def write_header(ws):
    for i, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = header_font; c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws.row_dimensions[1].height = 26

def section(name):
    r = ws.max_row + 1
    ws.cell(row=r, column=1, value=name).font = section_font
    ws.cell(row=r, column=1).fill = section_fill
    for col in range(2, 7):
        ws.cell(row=r, column=col).fill = section_fill
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[r].height = 22

def subsection(name):
    r = ws.max_row + 1
    ws.cell(row=r, column=1, value=name).font = sub_font
    for col in range(1, 7):
        ws.cell(row=r, column=col).fill = sub_fill
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=2)

def row(*vals, note=False):
    r = ws.max_row + 1
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=i, value=v if v is not None else "")
        c.font = body_font
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = border
        if note and i == 5:
            c.fill = note_fill
    ws.row_dimensions[r].height = 30

write_header(ws)

# ============================================================
# PÁGINA 1 — RESUMO CRM
# ============================================================
section("PÁGINA 1 — RESUMO CRM")

subsection("KPIs Topline · CRM Total")
row("Faturamento CRM", "Função", "Soma do faturamento dos 5 canais (atribuição last-click via UTM)", "Por dia / Por mês", "Supabase (consolida Shopify + UTM)", "KPI principal de gestão")
row("Compras", "Função", "Soma de compras dos 5 canais (last-click)", "Por dia / Por mês", "Supabase (consolida Shopify)", "Acompanhar volume")
row("Ticket Médio", "Função", "Faturamento / Compras", "Por dia / Por mês", "Supabase", "Valor médio do pedido")
row("RPS (Revenue per Session)", "Função", "Faturamento / Sessões  (= TKM × Conversão)", "Por dia / Por mês", "Supabase", "Eficiência geral do site CRM")
row("Conversão Geral", "Função", "Compras / Sessões", "Por dia / Por mês", "Supabase", "Eficiência do funil")

subsection("Receita CRM segmentada · 1ª Compra vs Recompra")
row("Receita 1ª Compra", "Função", "Soma faturamento de pedidos com customer.orders_count = 1", "Por canal / Por dia", "Shopify (orders + customers)", "Aquisição via CRM")
row("Receita Recompra", "Função", "Soma faturamento de pedidos com customer.orders_count > 1", "Por canal / Por dia", "Shopify (orders + customers)", "Retenção via CRM")
row("Receita Total CRM (MTD)", "Função", "Receita 1ª Compra + Receita Recompra (acumulado do mês)", "Consolidado", "Supabase", "Visão executiva")
row("Meta 1ª Compra", "Dado", None, "Mensal", "Input mensal (planilha de metas)", "Acompanhamento de meta")
row("Meta Recompra", "Dado", None, "Mensal", "Input mensal", "Acompanhamento de meta")
row("Meta Total CRM", "Dado", None, "Mensal", "Input mensal", "Acompanhamento de meta")
row("Meta x Realizado (1ª / Recompra / Total)", "Função", "Realizado / Meta", "Mensal", "Supabase", "Pace do mês")

subsection("Faturamento por canal")
row("Receita por canal (5 canais)", "Função", "Soma faturamento atribuído last-click via UTM", "Por canal / Por dia", "Supabase (Shopify + UTM)", "Ranking de canais")
row("% do total por canal", "Função", "Receita do canal / Receita CRM Total", "Por canal", "Supabase", "Share de cada canal")

subsection("Indicadores de Eficiência por canal (Resumo)")
row("Sessões (por canal)", "Dado", None, "Por canal / Por dia", "Looker (GA4) — filtrado por UTM source/medium", "Eficiência do site")
row("Conversão Geral (por canal)", "Função", "Compras / Sessões", "Por canal / Por dia", "Supabase", "Eficiência do funil")
row("Ticket Médio (por canal)", "Função", "Faturamento / Compras", "Por canal / Por dia", "Supabase", "Valor médio por canal")
row("RPS (por canal)", "Função", "Faturamento / Sessões", "Por canal / Por dia", "Supabase", "Eficiência total por canal")

subsection("Captação de Leads")
row("Leads novos por dia (total)", "Função", "Soma de novos profiles criados (e-mail + WhatsApp)", "Por dia / Por fonte", "Klaviyo (Profiles API)", "Acompanhar entrada de leads")
row("Leads médios por dia", "Função", "Leads MTD / dias do mês", "Mensal", "Supabase", "Saúde da captação")
row("Pico diário de leads", "Função", "MAX(Leads por dia) no período", "Mensal", "Supabase", "Detecção de picos / anomalias")
row("Taxa de Cadastro média", "Função", "Total Inscritos / Total Visualizações", "Mensal", "Supabase", "Eficiência geral de captação")
row("Visualizações totais (forms+popups)", "Função", "Soma de views de todos os ativos de captação", "Por dia / Por ativo", "Klaviyo Forms (Form Reports API)", "Base para cálculo de taxa")
row("Leads novos por fonte", "Dado", None, "Por fonte / Por dia", "Klaviyo (Profiles + source attribution)", "Identificar fontes mais produtivas")

subsection("Formulários e Popups · Performance de captação")
row("Nome do ativo de captação", "Dado", None, "Por ativo", "Klaviyo Forms", "Identificação (incluir local no próprio nome — ex.: 'Popup Home', 'Form Footer', 'Landing Coleção SS25')")
row("Tipo (Popup / Formulário / Embed)", "Dado", None, "Por ativo", "Klaviyo Forms", "Categorização")
row("Posicionamento", "Dado", None, "Por ativo", "Klaviyo Forms (campo nome) — todas as Landing Pages estão na Shopify", "Análise por contexto")
row("Visualizações por ativo", "Dado", None, "Por ativo / Por dia", "Klaviyo Forms API", "Base de cálculo")
row("Inscritos por ativo", "Dado", None, "Por ativo / Por dia", "Klaviyo Forms API", "Performance absoluta")
row("Taxa de Cadastro por ativo", "Função", "Inscritos / Visualizações", "Por ativo", "Supabase", "Performance relativa")

subsection("Funil CRM Consolidado")
row("Sessões (CRM total)", "Dado", None, "Por dia", "Looker (GA4)", "Topo do funil")
row("Add to Cart (ATC)", "Dado", None, "Por dia", "Looker (GA4)", "Etapa intermediária")
row("Begin Check-out (BCO)", "Dado", None, "Por dia", "Looker (GA4)", "Etapa intermediária")
row("Compras", "Dado", None, "Por dia", "Shopify", "Resultado final")
row("Taxa Sessão → ATC", "Função", "ATC / Sessões", "Por dia", "Supabase", "Gargalo de interesse")
row("Taxa ATC → BCO", "Função", "BCO / ATC", "Por dia", "Supabase", "Gargalo de intenção")
row("Taxa BCO → Compra", "Função", "Compras / BCO", "Por dia", "Supabase", "Gargalo de checkout")

subsection("Pace e Metas (Mensal)")
row("% atingido do mês", "Função", "Receita MTD / Meta total CRM", "Mensal", "Supabase", "Gauge de pace")
row("Projeção de fechamento", "Função", "(Receita MTD / dia atual) × dias do mês", "Mensal", "Supabase", "Antecipar resultado")
row("Status do pace (acima/abaixo)", "Função", "Projeção vs Meta", "Mensal", "Supabase", "Alerta")

subsection("Saúde da Base · E-mail")
row("Taxa de Entrega (Delivery Rate)", "Função", "E-mails entregues / E-mails enviados", "Por canal (EF, EC)", "Klaviyo API (Campaign + Flow metrics)", "Saúde da base")
row("Bounce Rate (Hard + Soft)", "Função", "Bounces / Enviados", "Por canal", "Klaviyo API", "Saúde da base")
row("Spam Complaint Rate", "Função", "Reclamações / Entregues", "Por canal", "Klaviyo API", "Risco de blacklist")
row("Opt-out Rate (descadastros)", "Função", "Unsubscribes / Entregues", "Por canal", "Klaviyo API", "Queima de base")
row("Base Ativa (engajados 90d)", "Dado", None, "Total / Por canal", "Klaviyo (segmento 'Engaged 90d') via API", "Tamanho real da base")

subsection("Saúde da Base · WhatsApp")
row("REMOVIDO DO ESCOPO V1", "—", "Segunda prioridade. Deveria vir do Meta Business Manager (Quality Rating, delivery, blocks), mas não será forçado agora.", "—", "FORA DE ESCOPO — revisitar quando acesso ao Meta estiver pronto", "Não exibir no dashboard V1", note=True)

subsection("Filtros Globais (todas as páginas)")
row("Filtro Período", "Função", "Define janela (D-1 / 7d / MTD / Mês anterior / 30d / Custom)", "Todas páginas", "Frontend", "Interatividade")
row("Filtro Comparativo", "Função", "Dia vs Dia anterior  /  Mês vs Mês anterior", "Todas páginas", "Frontend", "Δ contextual")
row("Filtro Canal", "Função", "Filtra todas as métricas pelo canal selecionado", "Todas páginas", "Frontend", "Drill-down rápido")
row("Filtro Tipo de Cliente", "Função", "Todos / 1ª Compra / Recompra", "Todas páginas", "Shopify (customer.orders_count)", "Segmentação")
row("Filtro Fluxo/Campanha", "Função", "Filtra por ativo individual", "Páginas de canal", "Klaviyo / Vekta / Sendflow (conforme canal)", "Drill-down por ativo")

# ============================================================
# PÁGINA 2 — E-MAIL FLUXO
# ============================================================
section("PÁGINA 2 — E-MAIL FLUXO")

subsection("KPIs do canal (Topo)")
row("Faturamento (canal)", "Função", "Soma faturamento atribuído last-click a E-mail Fluxo", "Por dia / Por mês", "Supabase (Shopify + UTM)", "Receita do canal")
row("Compras (canal)", "Função", "Soma compras atribuídas last-click", "Por dia / Por mês", "Supabase", "Volume do canal")
row("Ticket Médio (canal)", "Função", "Faturamento / Compras", "Por dia / Por mês", "Supabase", "TKM do canal")
row("RPS (canal)", "Função", "Faturamento / Sessões", "Por dia / Por mês", "Supabase", "Eficiência do canal")
row("Conversão Geral (canal)", "Função", "Compras / Sessões", "Por dia / Por mês", "Supabase", "Eficiência do funil")
row("Receita / Inscrito", "Função", "Faturamento / Inscritos (base ativa)", "Por dia / Por mês", "Supabase (Klaviyo + Shopify)", "KPI principal de eficiência do canal")

subsection("Funil completo do site · atribuído ao canal (last-click)")
row("Sessões (E-mail Fluxo)", "Dado", None, "Por dia / Por fluxo / Por e-mail", "Looker (GA4 — UTM source=email medium=flow)", "Topo do funil")
row("Add to Cart", "Dado", None, "Por dia / Por fluxo / Por e-mail", "Looker (GA4)", "Etapa intermediária")
row("Begin Check-out", "Dado", None, "Por dia / Por fluxo / Por e-mail", "Looker (GA4)", "Etapa intermediária")
row("Compras", "Dado", None, "Por dia / Por fluxo / Por e-mail", "Shopify (com UTM)", "Resultado")
row("Conversões etapa a etapa", "Função", "Cálculo padrão de funil", "Por dia", "Supabase", "Gargalo")

subsection("Receita por Inscrito · Diário")
row("Receita / Inscrito diário", "Função", "Receita do dia / Inscritos ativos no dia", "Por dia", "Supabase", "Acompanhar tendência")

subsection("Métricas específicas · E-mail Fluxo")
row("Inscritos Totais (canal)", "Dado", None, "Por fluxo", "Klaviyo", "Tamanho da base")
row("Envios de e-mails", "Dado", None, "Por fluxo / Por e-mail", "Klaviyo", "Volume disparado")
row("E-mails Abertos", "Dado", None, "Por fluxo / Por e-mail", "Klaviyo", "Engajamento topo")
row("Taxa de Abertura", "Função", "Abertos / Enviados", "Por fluxo / Por e-mail", "Supabase", "Eficiência do subject")
row("Cliques nos e-mails", "Dado", None, "Por fluxo / Por e-mail", "Klaviyo", "Engajamento fundo")
row("CTOR (Click-to-Open Rate)", "Função", "Cliques / Abertos", "Por fluxo / Por e-mail", "Supabase", "Eficiência do conteúdo (renomeado de 'CTR')")
row("Receita / Envio", "Função", "Faturamento / Envios", "Por fluxo / Por e-mail", "Supabase", "Eficiência por disparo")

subsection("Saúde da Base · E-mail Fluxo")
row("Taxa de Entrega", "Função", "Entregues / Enviados", "Por fluxo", "Klaviyo API", "Saúde")
row("Bounce Rate", "Função", "Bounces / Enviados", "Por fluxo", "Klaviyo API", "Saúde")
row("Spam Complaint", "Função", "Spam / Entregues", "Por fluxo", "Klaviyo API", "Risco")
row("Opt-out", "Função", "Unsubscribes / Entregues", "Por fluxo", "Klaviyo API", "Queima")
row("Base Ativa 90d", "Dado", None, "Canal", "Klaviyo (segmento via API)", "Base útil")

subsection("Detalhamento · Por Fluxo (com drill-down para E-mails)")
row("Linha-pai: cada fluxo", "Função", "Agregação de todas as métricas por fluxo", "Por fluxo", "Supabase (Klaviyo + Shopify + GA4)", "Visão por ativo")
row("Linha-filho: cada e-mail do fluxo", "Função", "Mesmas métricas no nível e-mail individual", "Por e-mail", "Supabase", "Drill-down")
row("Subject Line do e-mail", "Dado", None, "Por e-mail", "Klaviyo", "Identificação do criativo")

subsection("Ranking de Ativos · E-mail Fluxo")
row("Top 5 melhores e piores · por Receita", "Função", "ORDER BY receita DESC/ASC LIMIT 5", "Por fluxo", "Supabase", "Identificar vencedores/perdedores")
row("Top 5 melhores e piores · por Receita/Inscrito", "Função", "ORDER BY (receita/inscritos) DESC/ASC LIMIT 5", "Por fluxo", "Supabase", "Eficiência relativa")

# ============================================================
# PÁGINA 3 — E-MAIL CAMPANHA
# ============================================================
section("PÁGINA 3 — E-MAIL CAMPANHA")

subsection("KPIs do canal (Topo)")
row("Faturamento (canal)", "Função", "Soma faturamento atribuído last-click a E-mail Campanha", "Por dia / Por mês", "Supabase (Shopify + UTM)", "Receita do canal")
row("Compras (canal)", "Função", "Soma compras", "Por dia / Por mês", "Supabase", "Volume")
row("Ticket Médio", "Função", "Faturamento / Compras", "Por dia / Por mês", "Supabase", "TKM")
row("RPS", "Função", "Faturamento / Sessões", "Por dia / Por mês", "Supabase", "Eficiência")
row("Conversão Geral", "Função", "Compras / Sessões", "Por dia / Por mês", "Supabase", "Eficiência funil")
row("Receita / Disparo", "Função", "Faturamento / Envios", "Por dia / Por mês", "Supabase (Klaviyo + Shopify)", "KPI principal do canal")

subsection("Funil completo do site · atribuído ao canal")
row("Sessões / ATC / BCO / Compras (campanha)", "Dado/Função", "Funil padrão", "Por dia / Por campanha", "Looker (UTM source=email medium=campaign) + Shopify", "Funil")

subsection("Receita por Disparo · Diário")
row("Receita / Disparo diário", "Função", "Receita do dia / Envios do dia", "Por dia", "Supabase", "Tendência")

subsection("Métricas específicas · E-mail Campanha")
row("Envios de e-mails", "Dado", None, "Por campanha", "Klaviyo", "Volume")
row("E-mails Abertos", "Dado", None, "Por campanha", "Klaviyo", "Engajamento")
row("Taxa de Abertura", "Função", "Abertos / Enviados", "Por campanha", "Supabase", "Subject")
row("Cliques", "Dado", None, "Por campanha", "Klaviyo", "Engajamento")
row("CTOR", "Função", "Cliques / Abertos", "Por campanha", "Supabase", "Conteúdo")
row("Receita / Disparo", "Função", "Receita / Envios", "Por campanha", "Supabase", "Eficiência")

subsection("Saúde da Base · E-mail Campanha")
row("Taxa de Entrega / Bounce / Spam / Opt-out / Base Ativa", "Função/Dado", "Mesmas regras de Saúde da Base do canal e-mail", "Por campanha", "Klaviyo API", "Saúde")

subsection("Detalhamento · Por Campanha (sem drill-down — campanha = e-mail)")
row("Linha por campanha", "Função", "Agregação completa por campanha individual", "Por campanha", "Supabase", "Cada campanha é o próprio e-mail")

subsection("Ranking de Ativos · E-mail Campanha")
row("Top 5 melhores e piores · por Receita", "Função", "ORDER BY receita DESC/ASC LIMIT 5", "Por campanha", "Supabase", "Vencedores/perdedores")
row("Top 5 melhores e piores · por Receita/Disparo", "Função", "ORDER BY (receita/envios) DESC/ASC LIMIT 5", "Por campanha", "Supabase", "Eficiência relativa")

# ============================================================
# PÁGINA 4 — WHATSAPP FLUXO
# ============================================================
section("PÁGINA 4 — WHATSAPP FLUXO")

subsection("KPIs do canal (Topo)")
row("Faturamento (canal)", "Função", "Soma faturamento atribuído last-click a WhatsApp Fluxo", "Por dia / Por mês", "Supabase (Shopify + UTM)", "Receita")
row("Compras (canal)", "Função", "Soma compras", "Por dia / Por mês", "Supabase", "Volume")
row("Ticket Médio", "Função", "Faturamento / Compras", "Por dia / Por mês", "Supabase", "TKM")
row("RPS", "Função", "Faturamento / Sessões", "Por dia / Por mês", "Supabase", "Eficiência")
row("Conversão Geral", "Função", "Compras / Sessões", "Por dia / Por mês", "Supabase", "Eficiência funil")
row("Receita / Inscrito", "Função", "Faturamento / Inscritos ativos", "Por dia / Por mês", "Supabase (Hubspot hoje → Klaviyo após migração + Shopify)", "KPI principal do canal")

subsection("Funil completo do site · atribuído ao canal")
row("Sessões / ATC / BCO / Compras", "Dado/Função", "Funil padrão", "Por dia / Por fluxo / Por template", "Looker (UTM source=whatsapp medium=flow) + Shopify", "Funil")

subsection("Receita por Inscrito · Diário")
row("Receita / Inscrito diário", "Função", "Receita do dia / Inscritos ativos do dia", "Por dia", "Supabase", "Tendência")

subsection("Métricas específicas · WhatsApp Fluxo")
row("Inscritos Totais", "Dado", None, "Por fluxo", "Hubspot (atual) → Klaviyo (migração planejada — passar a ser fonte oficial)", "Tamanho da base", note=True)
row("Disparos Totais", "Dado", None, "Por fluxo / Por template", "Vekta + n8n", "Volume")
row("Cobertura (Disparos / Inscritos)", "Função", "Disparos / Inscritos", "Por fluxo", "Supabase", "Cobertura da base")
row("Respostas (Acessar)", "Dado", None, "Por fluxo / Por template", "Vekta + n8n", "Engajamento")
row("Taxa de Resposta", "Função", "Respostas / Disparos", "Por fluxo / Por template", "Supabase", "Eficiência do template")
row("Destravar Objeção", "Dado", None, "Por fluxo / Por template", "Vekta + n8n", "Avanço de qualificação")
row("Leads passados ao Comercial", "Dado", None, "Por fluxo", "Vekta + n8n", "Handoff")
row("Retenção da IA (Taxa de Handoff)", "Função", "Leads ao Comercial / Respostas", "Por fluxo", "Supabase", "Eficiência da IA")
row("CTR das mensagens", "Função", "Respostas do Acessar / Sessões", "Por template", "Supabase", "Eficiência do copy")
row("Receita / Inscrito", "Função", "Receita / Inscritos", "Por fluxo", "Supabase", "Eficiência")
row("Receita / Disparo", "Função", "Receita / Disparos", "Por fluxo", "Supabase", "Eficiência alternativa")

subsection("Detalhamento · Por Fluxo (com drill-down para Templates)")
row("Linha-pai: cada fluxo", "Função", "Agregação completa por fluxo", "Por fluxo", "Supabase (Vekta + n8n + Hubspot/Klaviyo + GA4 + Shopify)", "Visão por ativo")
row("Linha-filho: cada template do fluxo", "Função", "Mesmas métricas no nível template", "Por template", "Supabase", "Drill-down")
row("Conteúdo do template", "Dado", None, "Por template", "Vekta", "Identificação do criativo")

subsection("Ranking de Ativos · WhatsApp Fluxo")
row("Top 5 melhores e piores · por Receita", "Função", "ORDER BY receita DESC/ASC LIMIT 5", "Por fluxo", "Supabase", "Vencedores/perdedores")
row("Top 5 melhores e piores · por Receita/Inscrito", "Função", "ORDER BY (receita/inscritos) DESC/ASC LIMIT 5", "Por fluxo", "Supabase", "Eficiência relativa")

# ============================================================
# PÁGINA 5 — WHATSAPP CAMPANHA
# ============================================================
section("PÁGINA 5 — WHATSAPP CAMPANHA")

subsection("KPIs do canal (Topo)")
row("Faturamento (canal)", "Função", "Soma faturamento atribuído last-click a WhatsApp Campanha", "Por dia / Por mês", "Supabase (Shopify + UTM)", "Receita")
row("Compras (canal)", "Função", "Soma compras", "Por dia / Por mês", "Supabase", "Volume")
row("Ticket Médio", "Função", "Faturamento / Compras", "Por dia / Por mês", "Supabase", "TKM")
row("RPS", "Função", "Faturamento / Sessões", "Por dia / Por mês", "Supabase", "Eficiência")
row("Conversão Geral", "Função", "Compras / Sessões", "Por dia / Por mês", "Supabase", "Eficiência")
row("Receita / Disparo", "Função", "Faturamento / Disparos", "Por dia / Por mês", "Supabase (Vekta + Shopify)", "KPI principal — sem conceito de 'inscritos' por campanha")

subsection("Funil completo do site · atribuído ao canal")
row("Sessões / ATC / BCO / Compras (campanha)", "Dado/Função", "Funil padrão", "Por dia / Por campanha", "Looker (UTM source=whatsapp medium=campaign) + Shopify", "Funil")

subsection("Receita por Disparo · Diário")
row("Receita / Disparo diário", "Função", "Receita do dia / Disparos do dia", "Por dia", "Supabase", "Tendência")

subsection("Métricas específicas · WhatsApp Campanha")
row("Disparos", "Dado", None, "Por campanha / Por template", "Vekta (Spread em uso hoje não é fonte confiável — migração para Vekta em andamento)", "Volume", note=True)
row("Respostas (Acessar)", "Dado", None, "Por campanha / Por template", "Vekta", "Engajamento")
row("Taxa de Resposta", "Função", "Respostas / Disparos", "Por campanha / Por template", "Supabase", "Eficiência")
row("Destravar Objeção", "Dado", None, "Por campanha", "Vekta", "Avanço de qualificação")
row("Leads passados ao Comercial", "Dado", None, "Por campanha", "Vekta + n8n", "Handoff")
row("Retenção da IA (Taxa de Handoff)", "Função", "Leads ao Comercial / Respostas", "Por campanha", "Supabase", "Eficiência")
row("CTR das mensagens", "Função", "Respostas / Sessões", "Por template", "Supabase", "Eficiência copy")
row("Receita / Disparo", "Função", "Receita / Disparos", "Por campanha", "Supabase", "Eficiência")

subsection("Detalhamento · Por Campanha (com drill-down para Template)")
row("Linha-pai: cada campanha", "Função", "Agregação completa por campanha", "Por campanha", "Supabase", "Visão")
row("Linha-filho: template/peça da campanha", "Função", "Mesmas métricas + conteúdo do template", "Por template", "Vekta", "Drill-down")

subsection("Ranking de Ativos · WhatsApp Campanha")
row("Top 5 melhores e piores · por Receita", "Função", "ORDER BY receita DESC/ASC LIMIT 5", "Por campanha", "Supabase", "Vencedores/perdedores")
row("Top 5 melhores e piores · por Receita/Disparo", "Função", "ORDER BY (receita/disparos) DESC/ASC LIMIT 5", "Por campanha", "Supabase", "Eficiência relativa")

# ============================================================
# PÁGINA 6 — COMUNIDADE WHATSAPP
# ============================================================
section("PÁGINA 6 — COMUNIDADE WHATSAPP")

subsection("KPIs do canal (Topo)")
row("Faturamento (canal)", "Função", "Soma faturamento atribuído last-click à Comunidade", "Por dia / Por mês", "Supabase (Shopify + UTM)", "Receita")
row("Compras (canal)", "Função", "Soma compras", "Por dia / Por mês", "Supabase", "Volume")
row("Ticket Médio", "Função", "Faturamento / Compras", "Por dia / Por mês", "Supabase", "TKM")
row("RPS", "Função", "Faturamento / Sessões", "Por dia / Por mês", "Supabase", "Eficiência")
row("Conversão Geral", "Função", "Compras / Sessões", "Por dia / Por mês", "Supabase", "Eficiência")
row("Receita / Disparo", "Função", "Faturamento / Disparos", "Por dia / Por mês", "Supabase (Sendflow + Shopify)", "KPI principal")

subsection("Funil completo do site · atribuído à Comunidade")
row("Sessões / ATC / BCO / Compras", "Dado/Função", "Funil padrão", "Por dia", "Looker (UTM source=community) + Shopify", "Funil")

subsection("Receita por Disparo · Diário")
row("Receita / Disparo diário", "Função", "Receita do dia / Disparos do dia", "Por dia", "Supabase", "Tendência")

subsection("Métricas específicas · Comunidade")
row("Participantes", "Dado", None, "Total / Diário", "Sendflow", "Tamanho")
row("Entradas (30d)", "Dado", None, "Diário", "Sendflow", "Crescimento")
row("Saídas (30d)", "Dado", None, "Diário", "Sendflow", "Churn")
row("Crescimento líquido", "Função", "Entradas - Saídas", "Mensal", "Supabase", "Saúde")
row("Disparos", "Dado", None, "Por mensagem", "Sendflow", "Volume")
row("Receita / Participante", "Função", "Receita / Participantes", "Mensal", "Supabase", "LTV simplificado")
row("Receita / Disparo", "Função", "Receita / Disparos", "Por mensagem", "Supabase", "Eficiência")

subsection("Detalhamento · Por Mensagem (lista direta)")
row("Linha por mensagem (sem drill-down)", "Função", "Cada disparo = um ativo", "Por mensagem", "Sendflow + Supabase", "Visão direta")

subsection("Ranking de Ativos · Comunidade")
row("Top 5 melhores e piores · por Receita", "Função", "ORDER BY receita DESC/ASC LIMIT 5", "Por mensagem", "Supabase", "Vencedores/perdedores")
row("Top 5 melhores e piores · por Receita/Disparo", "Função", "ORDER BY (receita/disparos) DESC/ASC LIMIT 5", "Por mensagem", "Supabase", "Eficiência")

# Configura larguras
widths = [42, 18, 50, 28, 44, 38]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"

# ============================================================
# ABA: NOTAS & PRÓXIMOS PASSOS
# ============================================================
ws2 = wb.create_sheet("Notas & Próximos Passos")
ws2.cell(row=1, column=1, value="Notas Arquiteturais & Próximos Passos").font = Font(name=FONT_NAME, bold=True, size=14)
ws2.row_dimensions[1].height = 26

notes = [
    ("Base de Dados", "Será criada no Supabase. Centraliza dados de Shopify, Klaviyo, Vekta, n8n, Sendflow, GA4/Looker, Hubspot."),
    ("Atribuição", "Last-click via UTM. Soma da receita dos 5 canais = Receita CRM Total."),
    ("UTMs", "Padronizadas mas complexas. Próximo passo: criar aba 'Sem Atribuição' onde o dashboard mostra receita com UTM não reconhecida, e o usuário classifica manualmente em qual canal/fluxo/campanha encaixar. NÃO incluir no V1."),
    ("Saúde da Base WhatsApp", "REMOVIDO do V1. Segunda prioridade. Quando acesso à Meta Business Manager (WhatsApp Business Platform) estiver pronto: puxar Quality Rating, delivery rate, blocks e opt-outs."),
    ("WhatsApp Campanha — Vekta", "Spread (em uso parcial hoje) não é fonte de dados confiável. Migração planejada para Vekta. Dashboard considera Vekta como fonte oficial daqui em diante."),
    ("WhatsApp Fluxo — Inscritos", "Hubspot é a fonte atual. Migração para Klaviyo planejada. Quando concluída, trocar fonte para Klaviyo sem mudar a métrica."),
    ("Landing Pages", "Todas hospedadas na Shopify. Não há ferramenta de LP separada. O nome do ativo de captação deve incluir o local (ex.: 'Popup Home', 'Form Footer', 'Landing Coleção SS25')."),
    ("Comunidade — Engajamento", "Não será mensurado no V1. Sendflow não expõe engajamento detalhado e não é prioridade."),
    ("CAC / LTV / ROAS", "Fora do escopo V1. Próxima página do dashboard quando custos por canal estiverem disponíveis."),
    ("Recompra / Coorte / Qualidade de Lead Comercial / Tempo de Ciclo", "Fora do escopo V1. Revisitar em fases futuras."),
    ("Cadência", "Dashboard de leitura diária. Operação opera por ele todo dia de manhã + gestão consome o resumo.")
]
for i, (titulo, conteudo) in enumerate(notes, 3):
    c1 = ws2.cell(row=i, column=1, value=titulo)
    c1.font = Font(name=FONT_NAME, bold=True, size=11)
    c1.alignment = Alignment(vertical="top", wrap_text=True)
    c2 = ws2.cell(row=i, column=2, value=conteudo)
    c2.font = Font(name=FONT_NAME, size=10)
    c2.alignment = Alignment(vertical="top", wrap_text=True)
    ws2.row_dimensions[i].height = 42
ws2.column_dimensions["A"].width = 38
ws2.column_dimensions["B"].width = 110

wb.save(r"C:\Users\danie\Downloads\Métricas Gerenciamento v2.xlsx")
print("OK")
